"""
BaseImporter — the strategy contract every import type implements.

Reading order per row:
    detect_columns (once)  -> extract_mapped -> normalize_row
    -> validate_row -> match -> upsert (commit only)

Only mapped source columns are pulled into ``cleaned_data``; every unmapped
column is dropped in ``extract_mapped`` and never enters staging.
"""
import csv
import io
import re
from decimal import Decimal, InvalidOperation

ZERO = Decimal("0")

# Currency symbols / stray marks stripped from money fields (৳ = Bengali Taka).
_CURRENCY = "৳$€£₹"
_WS_RE = re.compile(r"\s+")


class BaseImporter:
    import_type = ""
    TARGET_FIELDS = []          # list[Field] — order = mapping UI sequence
    MATCH_KEY = None            # platform field name used to match live rows
    # How a repeated match key WITHIN one file is treated: products reject a
    # duplicate SKU ("error"); dues just overwrite the party ("warning").
    duplicate_in_file = "error"

    def get_model(self):
        raise NotImplementedError

    def restore_snapshot(self, obj, prev):
        raise NotImplementedError

    def serialize(self, cleaned):
        """Make a normalized row JSON-safe for staging (Decimal → str)."""
        from decimal import Decimal as _D
        out = {}
        for f in self.TARGET_FIELDS:
            if f.name in cleaned:
                v = cleaned[f.name]
                out[f.name] = str(v) if isinstance(v, _D) else v
        return out

    def hydrate(self, stored):
        """Rebuild Decimals from a staged row before commit."""
        out = {}
        for f in self.TARGET_FIELDS:
            if f.name in stored and stored[f.name] is not None:
                if f.kind in ("money", "qty"):
                    out[f.name] = self.clean_number(stored[f.name])
                else:
                    out[f.name] = stored[f.name]
        return out

    # ---- column detection ---------------------------------------------------
    @staticmethod
    def read_table(django_file, filename=""):
        """Return (headers, rows) from a CSV or XLSX upload.

        ``headers`` is a list[str] (empty string for blank cells). ``rows`` is a
        list[list[str]]. Header row is assumed to be the first row; a headerless
        file simply yields positional headers upstream.
        """
        name = (filename or getattr(django_file, "name", "") or "").lower()
        django_file.seek(0)
        if name.endswith(".xlsx"):
            return BaseImporter._read_xlsx(django_file)
        return BaseImporter._read_csv(django_file)

    @staticmethod
    def _read_csv(django_file):
        data = django_file.read()
        if isinstance(data, bytes):
            data = data.decode("utf-8-sig", errors="ignore")
        reader = csv.reader(io.StringIO(data))
        rows = [[(c or "").strip() for c in row] for row in reader if any((c or "").strip() for c in row)]
        if not rows:
            return [], []
        return rows[0], rows[1:]

    @staticmethod
    def _read_xlsx(django_file):
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(django_file.read()), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for r in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c).strip() for c in r]
            if any(cells):
                rows.append(cells)
        wb.close()
        if not rows:
            return [], []
        return rows[0], rows[1:]

    def detect_columns(self, django_file, filename=""):
        headers, _ = self.read_table(django_file, filename)
        cols = []
        for i, h in enumerate(headers):
            cols.append({"index": i, "header": h or f"Column {i + 1}"})
        return cols

    def suggest_mapping(self, detected_columns):
        """Fuzzy auto-match each target field to a detected column by header.
        Returns {platform_field: index}. Admin can override every choice."""
        norm = lambda s: re.sub(r"[^a-z0-9]", "", (s or "").lower())
        by_norm = {}
        for col in detected_columns:
            by_norm.setdefault(norm(col["header"]), col["index"])
        out = {}
        for f in self.TARGET_FIELDS:
            for cand in (f.name, f.label):
                key = norm(cand)
                if key and key in by_norm:
                    out[f.name] = by_norm[key]
                    break
        return out

    # ---- per-row pipeline ---------------------------------------------------
    def extract_mapped(self, row, column_mapping):
        """Pull ONLY the mapped source columns into platform fields, in
        TARGET_FIELDS order. Unmapped columns are dropped here."""
        out = {}
        for f in self.TARGET_FIELDS:
            src = column_mapping.get(f.name)
            if src is None or src == "":
                continue
            idx = int(src)
            out[f.name] = row[idx] if 0 <= idx < len(row) else ""
        return out

    def normalize_row(self, mapped):
        cleaned = {}
        for f in self.TARGET_FIELDS:
            if f.name not in mapped:
                continue
            raw = mapped[f.name]
            if f.kind in ("money", "qty"):
                cleaned[f.name] = self.clean_number(raw)
            else:
                cleaned[f.name] = self.clean_text(raw)
        return cleaned

    def validate_row(self, cleaned):
        """Base checks: required present, numbers parseable. Subclasses extend
        with warnings via ``extra_validate``. Returns (status, errors)."""
        errors = []
        for f in self.TARGET_FIELDS:
            val = cleaned.get(f.name)
            if f.required and (val is None or val == ""):
                errors.append(f"{f.label} is required")
            if f.kind in ("money", "qty") and f.name in cleaned and cleaned[f.name] is None:
                errors.append(f"{f.label} is not a valid number")
        if errors:
            return "error", errors
        warnings = self.extra_validate(cleaned)
        return ("warning" if warnings else "valid"), warnings

    def extra_validate(self, cleaned):
        return []

    def match(self, cleaned, shop):
        raise NotImplementedError

    def upsert(self, cleaned, existing, shop, job):
        raise NotImplementedError

    def match_value(self, cleaned):
        """Dedupe key for within-file duplicate detection."""
        return (cleaned.get(self.MATCH_KEY) or "").strip().upper()

    # ---- normalization helpers ---------------------------------------------
    @staticmethod
    def clean_text(value):
        if value is None:
            return ""
        s = _WS_RE.sub(" ", str(value)).strip()
        return "" if s.lower() == "none" else s

    @staticmethod
    def clean_number(value):
        """Strip currency symbols, spaces and thousand separators (both Western
        ``6,210.00`` and Bengali lakh ``2,96,300.00`` reduce correctly once all
        commas are removed). Returns Decimal or None if unparseable/empty."""
        if value is None:
            return None
        s = str(value).strip()
        for ch in _CURRENCY:
            s = s.replace(ch, "")
        s = s.replace(",", "").replace(" ", "")
        if s in ("", "-"):
            return None
        try:
            return Decimal(s)
        except (InvalidOperation, ValueError):
            return None

    @staticmethod
    def norm_phone(value):
        """Digits only — used for party matching, not storage."""
        return re.sub(r"\D", "", value or "")

    @staticmethod
    def _assert_tenant(obj, shop):
        assert obj.shop_id == shop.id, (
            f"tenant leak: object shop {obj.shop_id} != job shop {shop.id}"
        )
